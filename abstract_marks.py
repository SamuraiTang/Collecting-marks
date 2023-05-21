import openpyxl
import os
# put this python script in the directory of the file containing the progress test results,
# build the Summary.xlsx and past the student id in the firls column from the marking sheet

lec_code = 'EEEE2058'
aim_stu = 49

num_list = '0123456789'
file_xlsx = []
for filename in os.listdir(r'.'):
    name_select = filename[0:4]
    if name_select == 'Mark':
        file_xlsx.append(filename)

wb_sum = openpyxl.load_workbook('Summary.xlsx')
sh_sum = wb_sum['Sheet1']

for k in range(len(file_xlsx)):
    wb = openpyxl.load_workbook(file_xlsx[k], data_only=True)
    sh = wb['Marks']
    first_row = [row for row in sh.rows][0]
    first_row_items = [cell.value for cell in first_row]

    # sometimes the columns "Student ID" and "Full Name" are misplaced
    temp1 = first_row_items.index('Student ID') + 1
    temp2 = sh.cell(row=2, column=temp1).value
    if type(temp2) is str:
        if temp2[0] in num_list:
            original_stu_col = temp1
        else:
            original_stu_col = first_row_items.index('Full Name') + 1
    else:
        original_stu_col = temp1

    temp1 = [sh.cell(row=j+2, column=1).value for j in range(sh.max_row)]
    temp2 = [temp1[j] for j in range(len(temp1)) if temp1[j] is not None]
    original_stu = len(temp2)

    original_stu_marks_col = first_row_items.index(lec_code) + 1

    original_data = []
    original_marks = []

    for i in range(original_stu):
        original_data.append(int(sh.cell(row=i + 2, column=original_stu_col).value))
        temp = sh.cell(row=i + 2, column=original_stu_marks_col).value
        if temp is not None:
            if temp == 'EC':
                original_marks.append(8888)  # EC for additional operation
            else:
                original_marks.append(int(sh.cell(row=i + 2, column=original_stu_marks_col).value))
        else:
            original_marks.append(int(0))

    for i in range(aim_stu):
        aim_id = int(sh_sum.cell(row=i+1, column=1).value)
        temp = original_data.index(aim_id)
        sh_sum.cell(row=i + 1, column=k + 2).value = original_marks[temp]

wb_sum.save('Summary.xlsx')
